# import s3fs
import json
import pandas as pd  
from openpyxl import load_workbook
from distutils.util import strtobool
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


# fs = s3fs.S3FileSystem(use_listings_cache=False)

def load_json(json_path):
	# if(not fs.exists(json_path)):
	# 	return {}
	content = ""
	# with fs.open(json_path, 'rb') as f_json:
	with open(json_path, 'rb') as f_json:
		content += f_json.read().decode('utf-8')
	return json.loads(content)

def get_coordinate_index(cell_coordinate):
	xy = coordinate_from_string(cell_coordinate)
	col = column_index_from_string(xy[0])
	row = xy[1]

	return col, row

def get_filter_index(ws):
	init_col_filter = None
	init_row_filter = None
	for row in ws.rows:
		for cell in row:
			if(cell.value == "Filter criteria"):
				init_col_filter, init_row_filter = get_coordinate_index(cell.coordinate)
				init_row_filter += 1
	
	return init_col_filter, init_row_filter

def get_date_range_index(ws):
	for row in ws.rows:
		for cell in row:
			if(cell.value == "Date Range"):
				init_col_date, init_row_date = get_coordinate_index(cell.coordinate)
				init_col_date += 1
				return init_row_date, init_col_date

def get_final_index(ws, init_row_id, init_col_id):
	for row in ws.iter_rows(min_row=init_row_id, min_col=init_col_id, max_col=init_col_id):
		for cell in row:
			if(cell.value != None):
				_, last_row_id = get_coordinate_index(cell.coordinate)
			else:
				return last_row_id

def get_content(init_row_id, last_row_id, init_col_id, last_col_id):
	data = []
	for row in ws.iter_rows(min_row=init_row_id, max_row=last_row_id, min_col=init_col_id, max_col=last_col_id):
		rw = []
		for cell in row:
			rw.append(cell.value)
		data.append(rw)
	
	return data

def melt_df(df, melt_parameters):
	return df.melt(id_vars=melt_parameters["id_vars"],
		var_name=melt_parameters["var_name"], 
		value_name=melt_parameters["value_name"])

def add_date_range(data, date_range_row, date_range_col):
	new_data = []
	for row in data:
		rw = []
		rw.append(ws.cell(date_range_row, date_range_col).value)
		for cell in row:
			rw.append(cell)
		new_data.append(rw)
	
	return new_data


parameters = load_json("/home/mauricio/workspace/python-xls-to-csv/sheets_parameters.json")
wb = load_workbook("/home/mauricio/Downloads/socialbakers.xlsx")

for sheet in parameters:
	filter_concat = []
	ws = wb[wb.sheetnames[sheet["id"]]]
	cols = sheet["cols"]

	date_range_row, date_range_col = get_date_range_index(ws)

	init_col_id, init_row_id = get_coordinate_index(sheet["table_init_cell"])
	last_col_id = len(cols)
	last_row_id = get_final_index(ws, init_row_id, init_col_id)

	data = get_content(init_row_id, last_row_id, init_col_id, last_col_id)
	df_melted = pd.DataFrame(data, columns=cols)

	if(strtobool(sheet["melt_df"])):
		melt_parameters = sheet["melt_parameters"]
		cols = melt_parameters["id_vars"] + \
			[melt_parameters["var_name"]] + \
			[melt_parameters["value_name"]]
		df_melted = melt_df(df_melted, melt_parameters)	

	data = df_melted.values.tolist()
	data = add_date_range(data, date_range_row, date_range_col)

	init_col_filter, init_row_filter = get_filter_index(ws)
	print(init_col_filter)
	if(init_col_filter is not None):
		last_col_filter = init_col_filter + 2
		last_row_filter = get_final_index(ws, init_row_filter, init_col_filter)

		filters = get_content(init_row_filter, last_row_filter, init_col_filter, last_col_filter)
	else:
		filters = None

	if(filters is None):
		filter_concat.append(None)
		filter_concat.append(None)
		filter_concat.append(None)
		filter_concat.append(None)
	elif(len(filters) > 1):
		for filter in filters:
			filter.pop(0)
			filter_concat.append(filter[0])
			filter_concat.append(filter[1])
	else:
		for filter in filters:
			filter.pop(0)
			filter_concat.append(filter[0])
			filter_concat.append(filter[1])
			filter_concat.append(None)
			filter_concat.append(None)

	for row in data:
		for item in filter_concat:
			row.append(item)
	
	cols = ["date_range"] + cols + ["filter_1_type", "filter_1_value", "filter_2_type", "filter_2_value"]
	df = pd.DataFrame(data, columns=cols)
	df.to_csv(f"/home/mauricio/workspace/python-xls-to-csv/output/{sheet['name']}.csv", index=False, sep=";")
